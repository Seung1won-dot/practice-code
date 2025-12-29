from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "데이터"

ws['A1'] = "이름"
ws['B1'] = "나이"
ws['C1'] = "점수"

data = [
    ["홍길동", 25, 85],
    ["김철수", 30, 92],
    ["이영희", 28, 78]
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save("example.xlsx")

wb2 = load_workbook("example.xlsx")
ws2 = wb2.active

for row in ws2.iter_rows(min_row=2, values_only=True):
    print(f"이름: {row[0]}, 나이: {row[1]}, 점수: {row[2]}")

ws2['D1'] = "평균"
for row in range(2, ws2.max_row + 1):
    score = ws2.cell(row=row, column=3).value
    ws2.cell(row=row, column=4, value=score)

wb2.save("example_updated.xlsx")

